""""
Crawler : 웹 브라우저를 흉내내는 자동화 프로그램
1. 원하는 웹페이지에 접속해서 소스코드(HTML) 받아오기
- requests
$ pip install requests
2. 제대로 받아졌는지 확인(status code 확인)
3. 받아온 소스코드를 해석할 수 있는 형태로 파싱(해석, 변환)
-BeautifulSoup4
$pip install BeautifulSoup4

4. 해당 데이터에 원하는 요소 찾기

* 1. 셀레늄 - 보안 문자 직접 입력
* 2. 머신러닝 기법 사용 - 보안 문자 학습하기
"""
url = "https://www.naver.com/"

import requests
# get, post, put

data = requests.get(url)

if data.status_code != requests.codes.ok:
    print("접속 실패")
    exit()

print(data.text)

from bs4 import  BeautifulSoup
html = BeautifulSoup(data.text, "html.parser")
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
tags = html.select(".PM_CL_realtimeKeyword_list_base .ah_a")
ws.cell(1, 1, "순위")
ws.cell(1, 2, "검색어")
ws.cell(1, 3, "링크")

for index, tag in enumerate(tags, start=2):
    rank = tag.select_one('.ah_r').text
    keyword = tag.select_one('.ah_k').text

    ws.cell(index, 1, rank)
    ws.cell(index, 2, keyword)
    ws.cell(index, 3, tag.attrs['href'])

    wb.save("naver_keyword.xlsx")
    print(rank, keyword, tag.attrs['href'])

# data save in excel file
# html.select_one("selector")

"""
css selector
1. tag : sapn
2. id : #sapn
3. class : .span

복합 셀렉터
1. tag .other_class  얘는 중간 경로 안 써도 됨
2. tag > .other_class 얘는 중간 경로 다 써야 됨
"""